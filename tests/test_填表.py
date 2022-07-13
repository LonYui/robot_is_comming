"""默認=happycase"""
import datetime


def test_整合測試_填表():
    """填寫更加好報表"""
    # wb = 執行主程式()
    from openpyxl import Workbook
    更加好報表_機器人_wb = Workbook()
    更加好報表_機器人 = 更加好報表_機器人_wb.active
    # 更加好報表_機器人 = 執行主程式()
    """驗證填寫前後"""
    from openpyxl import load_workbook
    更加好報表_手打_wb = load_workbook('更加好_銷售報表07.03拷貝.xlsx')
    更加好報表_手打 = 更加好報表_手打_wb.active

    """日期正確"""
    更加好報表_手打_日期rows = []
    for row in 更加好報表_手打.iter_rows(min_row=8, max_row=10, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_手打_日期rows.append(row)
    更加好報表_機器人_日期rows = []
    for row in 更加好報表_機器人.iter_rows(min_row=8, max_row=10, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_機器人_日期rows.append(row)
    assert 更加好報表_機器人_日期rows == 更加好報表_手打_日期rows

    """日銷售正確"""
    更加好報表_手打_日銷售row = []
    for row in 更加好報表_手打.iter_rows(min_row=11, max_row=11, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_手打_日銷售row.append(row)
    更加好報表_機器人_日銷售row = []
    for row in 更加好報表_機器人.iter_rows(min_row=11, max_row=11, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_機器人_日銷售row.append(row)
    assert 更加好報表_機器人_日銷售row==更加好報表_手打_日銷售row

    """周銷售正確"""
    更加好報表_手打_週銷售row=[]
    for row in 更加好報表_手打.iter_rows(min_row=13, max_row=13, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_手打_週銷售row.append(row)
    更加好報表_機器人_週銷售row=[]
    for row in 更加好報表_機器人.iter_rows(min_row=13, max_row=13, min_col=3,
                               max_col=3 + 41, values_only=True):
        更加好報表_機器人_週銷售row.append(row)

    assert 更加好報表_機器人_週銷售row == 更加好報表_手打_週銷售row

def test_整合測試_填表2():
    '''填寫啵獅報表'''

    from openpyxl import Workbook
    啵獅報表_機器人_wb = Workbook()
    啵獅報表_機器人 = 啵獅報表_機器人_wb.active
    # 啵獅報表_機器人 = 執行主程式()
    """驗證填寫前後"""
    from openpyxl import load_workbook
    try:
        啵獅報表_手打_wb = load_workbook('7月/啵獅_銷售報表07.03拷貝.xlsx')
    except :
        啵獅報表_手打_wb = load_workbook('tests/7月/啵獅_銷售報表07.03.xlsx')
    啵獅報表_手打 = 啵獅報表_手打_wb.active
    
    """日期正確"""
    啵獅_日期cell位置 = {'min_row':10,'max_row':12,'min_col':3,'max_col':3+41}
    啵獅報表_手打_日期rows = []
    for row in 啵獅報表_手打.iter_rows(**啵獅_日期cell位置, values_only=True):
        啵獅報表_手打_日期rows.append(row)
    啵獅報表_機器人_日期rows = []
    for row in 啵獅報表_機器人.iter_rows(**啵獅_日期cell位置, values_only=True):
        啵獅報表_機器人_日期rows.append(row)
    assert 啵獅報表_機器人_日期rows == 啵獅報表_手打_日期rows

    """周銷售正確"""
    啵獅_商品cell位置 = {'min_row': 13, 'max_row': 28, 'min_col': 3,
                   'max_col': 3 + 41}
    啵獅報表_手打_週銷售row=[]
    for row in 啵獅報表_手打.iter_rows(**啵獅_商品cell位置, values_only=True):
        啵獅報表_手打_週銷售row.append(row)
    啵獅報表_機器人_週銷售row=[]
    for row in 啵獅報表_機器人.iter_rows(**啵獅_商品cell位置, values_only=True):
        啵獅報表_機器人_週銷售row.append(row)

    assert 啵獅報表_機器人_週銷售row == 啵獅報表_手打_週銷售row

    """商品價格正確"""
    assert 啵獅報表_手打[f'B{15+4*0}'].value == 啵獅報表_機器人[f'B{15+4*0}'].value
    assert 啵獅報表_手打[f'B{15+4*1}'].value == 啵獅報表_機器人[f'B{15+4*1}'].value
    assert 啵獅報表_手打[f'B{15+4*2}'].value == 啵獅報表_機器人[f'B{15+4*2}'].value
    assert 啵獅報表_手打[f'B{15+4*3}'].value == 啵獅報表_機器人[f'B{15+4*3}'].value

    '''檔期正確'''
    import string
    assert 啵獅報表_手打[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*0]}4'].value == 啵獅報表_機器人[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*0]}4'].value
    assert 啵獅報表_手打[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*1]}4'].value == 啵獅報表_機器人[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*1]}4'].value
    assert 啵獅報表_手打[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*2]}4'].value == 啵獅報表_機器人[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*2]}4'].value
    assert 啵獅報表_手打[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*3]}4'].value == 啵獅報表_機器人[f'{string.ascii_uppercase[string.ascii_uppercase.index("C")+7*3]}4'].value

    """title 正確"""
    啵獅報表_手打['A1'].value = 啵獅報表_機器人['A1'].value
    啵獅報表_手打['A2'].value = 啵獅報表_機器人['A2'].value
    啵獅報表_手打['A3'].value = 啵獅報表_機器人['A3'].value


def test_整合測試_更新config():
    """！這是手工
    進去google drive 填寫 config
    點擊發動，讓機器人跑一遍
    確認報表如預期產生
    """
    assert True

def test_讀取gooleDrive上config():
    """call ,確認和 config 同構"""
    下載的設定json = None
    from 填表 import 讀取googleDrive上的config
    下載的設定json= 讀取googleDrive上的config()
    import json
    設定_f = open('設定.json')
    設定 = json.load(設定_f)
    assert 設定[0].keys() == 下載的設定json[0].keys()


def test_get日銷售():
    '''
    查詢 7/8 donwon罐頭 銷售量
    '''
    # dongwon韓國東遠泡/菜鮪魚罐頭 7/8 銷售量是 43 jeffrey 人工查表 7/9
    input_商品代號_日期 = {'商品代號':'01040053','日期':datetime.date(2022,7,8)}
    ouput_銷售量 = 43

    銷售量 = 0
    # 銷售量 = get日銷售(**input_商品代號_日期)
    assert ouput_銷售量 == 銷售量

def test_get週銷售():
    '''
    查詢 七月第一週 picknic綠咖喱蝦風/味杯麵 銷售量
    '''
    # picknic綠咖喱蝦風/味杯麵 7月 week 1 銷售量是 300 + 300 = 600
    input_商品代號_日期 = {'商品代號': '01040053',
                     '資料年月': datetime.date(2022,7,1),'第幾週':1}
    output_週銷售量 = 300 + 300
    週銷售量 = 0
    # 週銷售量 = get銷售量(資料年月,第幾週,商品代號)
    assert output_週銷售量 == 週銷售量


def test_get尚存():
    '''blocked！ by 不能讀取每日的尚存
    查詢 七月 進銷存
    '''

# def test_上傳_報表到_gooleDrive():
#     '''
#     brief:上傳 excel 至 google drive
#     steps:製作一張excel 名稱 更加好_銷售報表<today()>
#     以此file為input 上傳_報表到_gooleDrive()
#     call api get files 路徑圖
#     assert 路徑格式 /更加好/latest 應該存在
#     assert 路徑格式 /更加好/7月w1/更加好_銷售報表07.03 應該存在
#     '''
#     from openpyxl import Workbook
#     更加好_銷售報表_today_wb = Workbook()
#     更加好_銷售報表_today_path = f'更加好_銷售報表{datetime.date.today().strftime("%m.%d")}.xlsx'
#     更加好_銷售報表_today_wb.save(更加好_銷售報表_today_path)
#
#     # 上傳_報表到_gooleDrive(更加好_銷售報表_today_path)
#
#     # call api
#
#
