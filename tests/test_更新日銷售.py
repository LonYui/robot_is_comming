'''更新日銷售 spec
作者:朧月
github:https://github.com/LonYui
'''
import os
import datetime


def test_更新日銷售():
    '''跑 7/3 更新，確定更加好有更新'''

    os.system('cp test_更新日銷售_測試資料/匯恩_銷售報表07.02.xlsx  db_test/匯恩_銷售報表07.xlsx') # 建立測試資料
    os.system('cp test_更新日銷售_測試資料/更加好_銷售報表07.02.xlsx  db_test/更加好_銷售報表07.xlsx')
    os.system('cp test_更新日銷售_測試資料/設定.json  db_test')

    from 更新日銷售 import main
    main()
    _比對_byrow('更加好',11)
    _比對_byrow('匯恩',11)
    _比對_byrow('匯恩',11+3)


    # 更加好報表_機器人_wb = load_workbook('tests/db_test/更加好_銷售報表07.xlsx')
    # 更加好報表_機器人 = 更加好報表_機器人_wb.active
    #
    # 更加好報表_手打_wb = load_workbook('tests/7月/更加好_銷售報表07.03.xlsx')
    # 更加好報表_手打 = 更加好報表_手打_wb.active
    #
    #
    # 更加好報表_手打_日銷售row = []
    # for row in 更加好報表_手打.iter_rows(min_row=11, max_row=11, min_col=3,
    #                               max_col=3 + 41, values_only=True):
    #     更加好報表_手打_日銷售row.append(row)
    # 更加好報表_機器人_日銷售row = []
    # for row in 更加好報表_機器人.iter_rows(min_row=11, max_row=11, min_col=3,
    #                                max_col=3 + 41, values_only=True):
    #     更加好報表_機器人_日銷售row.append(row)
    # assert 更加好報表_機器人_日銷售row == 更加好報表_手打_日銷售row    # 日銷售正確

    os.system('rm db_test/匯恩_銷售報表07.xlsx ')     # 刪除測試資料
    os.system('rm db_test/更加好_銷售報表07.xlsx ')     # 刪除測試資料
    os.system('rm db_test/設定.json ')

def test_ftp_每日銷售報表():
    '''下載本日檔案到db_test format:"219720220703sal.csv"'''
    input_outputs=[datetime.date.today().isoformat(),'2022-07-03']
    廠商編號s = [2197,2559]

    os.system('cp test_更新日銷售_測試資料/設定.json  db_test')

    from 更新日銷售 import ftp_每日銷售報表

    for input_output in input_outputs:
        ftp_每日銷售報表(input_output)
        日期_datetime = datetime.datetime.fromisoformat(input_output)
        for 廠商編號 in 廠商編號s:
            open(
                f'db_test/下載/{廠商編號}{日期_datetime.strftime("%Y%m%d")}sal.csv')
    # ftp_每日銷售報表()
    # for 廠商編號 in 廠商編號s:
    #     open(f'db_test/下載/{廠商編號}{datetime.datetime.today().strftime("%Y%m%d")}sal.csv')

    os.system('rm db_test/設定.json ')
    os.system('rm -r db_test/下載/')
    os.system('mkdir db_test/下載')



def test_get_日銷售():
    '''get 更加好 7/3 日銷售 from sal'''
    input_outputs=[({'貨號':'03030044','日期':'2022-07-03'},142)]  # 0.更加好 ,

    os.system('cp test_更新日銷售_測試資料/219720220703sal.csv db_test/下載/219720220703sal.csv') # 新增測試資料到ｄb
    os.system('cp test_更新日銷售_測試資料/設定.json  db_test')


    from 更新日銷售 import get_日銷售

    for input_output in input_outputs:
        input , output = input_output
        assert get_日銷售(**input) == output

    os.system('rm db_test/下載/219720220703sal.csv') # 刪除測試資料
    os.system('rm db_test/設定.json ')

def _比對_byrow(廠商, row):
    '''todo 可以比對多個商品 且只讀取一個 config'''
    from openpyxl import load_workbook  # 執行主程式()
    報表_機器人_wb = load_workbook(f'tests/db_test/{廠商}_銷售報表07.xlsx')
    報表_機器人 = 報表_機器人_wb.active
    報表_手打_wb = load_workbook(f'tests/7月/{廠商}_銷售報表07.03.xlsx')
    報表_手打 = 報表_手打_wb.active
    報表_手打_日銷售row = []
    for row in 報表_手打.iter_rows(min_row=row, max_row=row, min_col=3,
                               max_col=3 + 41, values_only=True):
        報表_手打_日銷售row.append(row)
    報表_機器人_日銷售row = []
    for row in 報表_機器人.iter_rows(min_row=row, max_row=row,
                                min_col=3,
                                max_col=3 + 41,
                                values_only=True):
        報表_機器人_日銷售row.append(row)
    assert 報表_機器人_日銷售row == 報表_手打_日銷售row  # 日銷售正確
